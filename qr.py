from openpyxl import load_workbook
import qrcode
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import time
import threading
import os

def open_xl():
    global xlText
    tp = [('Excelファイル','*.xlsx')]
    f = filedialog.askopenfilename(filetypes = tp)
    xlText.delete(0, tk.END)
    xlText.insert(tk.END, f)

def open_fld():
    global fldText
    f = filedialog.askdirectory()
    fldText.delete(0, tk.END)
    fldText.insert(tk.END, f)

def create_qrcode():
    th = threading.Thread(target=create_qrcode_thread)
    th.start()

def create_qrcode_thread():
    global xlText, fldText, outLabel, pb, pg
    try:
        wb = load_workbook(xlText.get())
    except:
        messagebox.showerror(u'エラー', u'ファイルを開くことができませんでした')
        return
    ws = wb.active
    r = 2
    while True:
        c = ws.cell(column = 1, row = r)
        if c.value is None:
            break
        r += 1
    maxRow = r - 2
#    print("maxRow = ", maxRow)
    fld = fldText.get()
    if fld == "":
        fld = os.path.dirname(xlText.get())
    if not os.path.isdir(fld):
        messagebox.showerror(u'エラー', u'出力先フォルダが存在しません')
        return
#    print(fld)
    r = 2
    while True:
        c = ws.cell(column = 1, row = r)
        if c.value is None:
            break
        c2 = ws.cell(column = 2, row = r)
        outText = u"出力中：" + c2.value + ".png"
        outLabel["text"] = outText
        img = qrcode.make(c.value)
        fname = os.path.join(fld, c2.value + '.png')
        try:
            img.save(fname)
        except:
            messagebox.showerror(u'エラー', fname + u'を保存することができませんでした')
        pg = (r - 1) / maxRow * 100
        pb.configure(value = pg)
        #print(r, c.value, c2.value)
        r += 1
        #time.sleep(1)
    outLabel["text"] = ''
    messagebox.showinfo(u'終了', u'QRコード作成が終了しました')

pg = 0
root = tk.Tk()
root.title(u"QRコード一括作成")
root.geometry("740x300")
xlLabel = tk.Label(text = u'Excelファイル', font=("MSゴシック", "16", "normal"))
xlLabel.place(x = 30, y = 30)
xlText = tk.Entry(width = 35, font=("MSゴシック", "16", "normal"))
xlText.place(x = 180, y = 30)
xlButton = tk.Button(text = u'参照', width = 10, font=("MSゴシック", "16", "normal"), command=open_xl)
xlButton.place(x = 590, y = 25)
fldLabel = tk.Label(text = u'出力先フォルダ', font=("MSゴシック", "16", "normal"))
fldLabel.place(x = 30, y = 80)
fldText = tk.Entry(width = 35, font=("MSゴシック", "16", "normal"))
fldText.place(x = 180, y = 80)
fldButton = tk.Button(text = u'参照', width = 10, font=("MSゴシック", "16", "normal"), command=open_fld)
fldButton.place(x = 590, y = 75)
makeButton = tk.Button(text = u'QRコード作成', width = 15, font=("MSゴシック", "20", "normal"), command=create_qrcode)
makeButton.place(x = 270, y = 130)
outLabel = tk.Label(text = u'', font=("MSゴシック", "16", "normal"))
outLabel.place(x = 30, y = 240)
pb = ttk.Progressbar(root, maximum=100, length = 680, mode="determinate")
pb.place(x = 30, y = 200)
root.mainloop()

